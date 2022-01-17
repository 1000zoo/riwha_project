from haversine import haversine
from openpyxl import Workbook
from openpyxl import load_workbook

class Point:
    def __init__ (self, latitude, longitude):
        self.latitude = latitude
        self.longitude = longitude

midExcel = Workbook()
m = midExcel.active

myFile = load_workbook('01-05.xlsx')
mF = myFile.active

lat = mF['A']
lng = mF['B']

midPoints = []
pointStack =[]

for i in range(len(lat) - 2):
    start = Point(lat[i+1].value, lng[i+1].value)
    goal = Point(lat[i+2].value, lng[i+2].value)
    
    pointStack.append(goal)
    current = start
    
    while not not pointStack:
        temp = pointStack[-1]
        s = (current.latitude, current.longitude)
        g = (temp.latitude, temp.longitude)
        if haversine(s, g, unit='m') >= 1:
            midLat = (current.latitude + temp.latitude) / 2
            midLon = (current.longitude + temp.longitude) / 2
            currentMidpoint = Point(midLat, midLon)
            pointStack.append(currentMidpoint)
        else:
            m.append([current.latitude, current.longitude])
            # midPoints.append(current)
            current = pointStack.pop()
            
    print(str(i) + " end")
    # midPoints.append(current)
    # print(midPoints)

midExcel.save('Midpoints.xlsx')