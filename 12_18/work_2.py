from haversine import haversine
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

myExcel = load_workbook('work2.xlsx')
myExcel.active
m = myExcel.worksheets[2]

result = Workbook()
midpoint_sheet = result.create_sheet('lane3_mid_point')
err_dist = result.create_sheet('err_dist')
lane3_mid_dist = result.create_sheet('lane3_mid_dist')

init_lat = m['A']
init_lng = m['B']
side_lat = m['D']
side_lng = m['E']
center_lat = m['G']
center_lng = m['H']
side_min = m['I']
lane_type = m['L']

for i in range(1, len(init_lat)):
    nlat = side_lat[i].value - abs(center_lat[i].value - side_lat[i].value) / 6
    nlng = side_lng[i].value - abs(center_lng[i].value - side_lng[i].value) / 6

    nmid = (nlat, nlng)
    cpoint = (center_lat[i].value, center_lng[i].value) #center_point
    ipoint = (init_lat[i].value, init_lng[i].value) #init_point

    lane3ToMid = haversine(nmid, cpoint, unit = 'm')
    midpoint_sheet.append([nlat, nlng])
    lane3_mid_dist.append([lane3ToMid])

    edist = haversine(cpoint, ipoint, unit = "m")

    if lane_type[i].value == "lane1" or lane_type[i].value == "lane2" or (lane_type[i].value == "lane3" and side_min[i].value > lane3ToMid):
        err_dist.append([-edist])
    else:
        err_dist.append([edist])

result.save('work2_result_sheet3.xlsx')
print("finish")


