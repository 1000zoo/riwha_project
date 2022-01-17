from haversine import haversine
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

myExcel = load_workbook('work1.xlsx')
m = myExcel.worksheets[0]

result = Workbook()
midpoint_sheet = result.create_sheet('lane1_mid_point')
err_dist = result.create_sheet('err_dist')
lane1_mid_dist = result.create_sheet('lane1_mid_dist')

init_lat = m['A']
init_lng = m['B']
side_lat = m['D']
side_lng = m['E']
center_lat = m['G']
center_lng = m['H']
cen_min = m['J']
lane_type = m['L']

for i in range(1, len(init_lat)):
    nlat = center_lat[i].value - abs(center_lat[i].value - side_lat[i].value) / 6
    nlng = center_lng[i].value + abs(center_lng[i].value - side_lng[i].value) / 6

    nmid = (nlat, nlng)
    cpoint = (center_lat[i].value, center_lng[i].value) #center_point
    ipoint = (init_lat[i].value, init_lng[i].value) #init_point

    lane1ToMid = haversine(nmid, cpoint, unit = 'm')
    midpoint_sheet.append([nlat, nlng])
    lane1_mid_dist.append([lane1ToMid])

    edist = haversine(nmid, ipoint, unit = "m")

    if lane_type[i].value == 'lane1':
        if edist > cen_min[i].value:
            err_dist.append([-edist])
        else:
            err_dist.append([edist])
    else:
        err_dist.append([edist])

result.save('work1_result_sheet1.xlsx')
print("finish")


