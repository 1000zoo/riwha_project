from haversine import haversine
from openpyxl import Workbook
from openpyxl import load_workbook

saveFile = Workbook()
s = saveFile.active

myFile = load_workbook("Midpoints.xlsx")
m = myFile.active
count = 0
lat = m['A']
lng = m['B']

for i in range(len(lat) - 1):
    start = (lat[i].value, lng[i].value)
    goal = (lat[i+1].value,lng[i+1].value)
    dist = haversine(start, goal, unit='m')
    print(dist)
    s.append([dist])
    count += 1

saveFile.save("distant.xlsx")
print(count)
