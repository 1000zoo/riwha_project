from haversine import haversine
from openpyxl import load_workbook
from openpyxl import Workbook

result1 = Workbook()
r1 = result1.active
result2 = Workbook()
r2 = result2.active

midpoint = load_workbook('lane_1211.xlsx')
m = midpoint.active

prilat = m['A']
prilng = m['B']
sidelat = m['D']
sidelng = m['E']
centerlat = m['G']
centerlng = m['H']
lane = m['L']
putMidpoint = m['N']
putDist = m['M']

for i in range(1, len(lane)):
    pri = (prilat[i].value, prilng[i].value)
    start = (sidelat[i].value, sidelng[i].value)
    goal = (centerlat[i].value, centerlng[i].value)
    mid = (round((start[0] + goal[0]) / 2, 5), round((start[1] + goal[1]) / 2, 5))
    print(pri)
    print(start)
    print(goal)
    dist = haversine(pri, mid, unit = 'm')

    if lane[i].value == "lane1":
        r1.append([-dist])
    else:
        r1.append([dist])
    
    r2.append([str(mid)])

result1.save('dist.xlsx')
result2.save('mid.xlsx')